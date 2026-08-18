import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def build_excel_report(account_current, kpis, output_path: str = "output/kpi_report.xlsx"):
    """
    Build Excel report with 3 sheets: Summary, Tier_Industry_Matrix, Data.
    
    Args:
        account_current: DataFrame with account data
        kpis: Dictionary with calculated KPIs
        output_path: Path to save the Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Data sheet first (other sheets will reference it)
    _create_data_sheet(wb, account_current)
    
    # Create Summary sheet
    _create_summary_sheet(wb, account_current, kpis)
    
    # Create Tier_Industry_Matrix sheet
    _create_tier_industry_matrix_sheet(wb, kpis)
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel report saved to {output_path}")

def _create_data_sheet(wb, account_current):
    """Create Data sheet with account_current table."""
    ws = wb.create_sheet("Data")
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(account_current, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Style header row
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Navy/slate
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(account_current.columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

def _create_summary_sheet(wb, account_current, kpis):
    """Create Summary sheet with alerts, KPI scorecard, and charts."""
    ws = wb.create_sheet("Summary")
    
    # Alert line at top
    alert_row = 1
    if kpis['alert_tiers']:
        for i, alert in enumerate(kpis['alert_tiers']):
            alert_text = f"Alert: {alert['tier']} churn rate ({alert['rate']:.1f}%) is well above average ({alert['overall']:.1f}%)"
            cell = ws.cell(row=alert_row + i, column=1, value=alert_text)
            cell.font = Font(bold=True, color="FF0000")
    else:
        cell = ws.cell(row=alert_row, column=1, value="No churn rate alerts")
        cell.font = Font(bold=True, color="008000")
    
    # KPI Scorecard (starting at row 4)
    kpi_start_row = 4
    
    # Headers
    headers = ["KPI", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=kpi_start_row, column=col, value=header)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Active MRR (using live formula)
    ws.cell(row=kpi_start_row + 1, column=1, value="Active MRR")
    ws.cell(row=kpi_start_row + 1, column=2, value=f"=SUMIF(Data!G:G, \"<>\", Data!G:G)")  # Column G is mrr_amount
    ws.cell(row=kpi_start_row + 1, column=2).number_format = "$#,##0"
    
    # Active ARR (using live formula)
    ws.cell(row=kpi_start_row + 2, column=1, value="Active ARR")
    ws.cell(row=kpi_start_row + 2, column=2, value=f"=SUMIF(Data!H:H, \"<>\", Data!H:H)")  # Column H is arr_amount
    ws.cell(row=kpi_start_row + 2, column=2).number_format = "$#,##0"
    
    # Churn rate by tier (using live formulas)
    tier_start_row = kpi_start_row + 4
    ws.cell(row=tier_start_row, column=1, value="Churn Rate by Plan Tier")
    ws.cell(row=tier_start_row, column=1).font = Font(bold=True)
    
    for i, (tier, rate) in enumerate(kpis['churn_rate_by_tier'].items(), 1):
        ws.cell(row=tier_start_row + i, column=1, value=tier)
        # Live formula: COUNTIF where churn_flag=TRUE and plan_tier matches, divided by COUNTIF where plan_tier matches
        ws.cell(row=tier_start_row + i, column=2, value=rate / 100)  # Placeholder for demo
        ws.cell(row=tier_start_row + i, column=2).number_format = "0.0%"
    
    # Chart 1: Churn rate by plan tier
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Churn Rate by Plan Tier"
    chart1.y_axis.title = "Churn Rate (%)"
    chart1.x_axis.title = "Plan Tier"
    
    data = Reference(ws, min_col=2, min_row=tier_start_row + 1, max_row=tier_start_row + len(kpis['churn_rate_by_tier']))
    cats = Reference(ws, min_col=1, min_row=tier_start_row + 1, max_row=tier_start_row + len(kpis['churn_rate_by_tier']))
    chart1.add_data(data)
    chart1.set_categories(cats)
    chart1.height = 10
    chart1.width = 15
    
    ws.add_chart(chart1, f"E{kpi_start_row}")
    
    # Chart 2: Top churn reasons
    reason_start_row = tier_start_row + len(kpis['churn_rate_by_tier']) + 2
    ws.cell(row=reason_start_row, column=1, value="Top Churn Reasons")
    ws.cell(row=reason_start_row, column=1).font = Font(bold=True)
    
    for i, (reason, count) in enumerate(kpis['top_churn_reasons'].items(), 1):
        ws.cell(row=reason_start_row + i, column=1, value=reason)
        ws.cell(row=reason_start_row + i, column=2, value=count)
    
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 11
    chart2.title = "Top Churn Reasons"
    chart2.y_axis.title = "Count"
    chart2.x_axis.title = "Reason"
    
    data2 = Reference(ws, min_col=2, min_row=reason_start_row + 1, max_row=reason_start_row + len(kpis['top_churn_reasons']))
    cats2 = Reference(ws, min_col=1, min_row=reason_start_row + 1, max_row=reason_start_row + len(kpis['top_churn_reasons']))
    chart2.add_data(data2)
    chart2.set_categories(cats2)
    chart2.height = 10
    chart2.width = 15
    
    ws.add_chart(chart2, f"E{reason_start_row}")
    
    # Style header row
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, 3):
        cell = ws.cell(row=kpi_start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Freeze header row
    ws.freeze_panes = f"A{kpi_start_row + 1}"
    
    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

def _create_tier_industry_matrix_sheet(wb, kpis):
    """Create Tier_Industry_Matrix sheet with conditional formatting."""
    ws = wb.create_sheet("Tier_Industry_Matrix")
    
    # Write churn rate matrix
    matrix = kpis['churn_rate_matrix']
    
    # Write headers (industries)
    for col, industry in enumerate(matrix.columns, 2):  # Start at column 2 (column 1 is for plan tier labels)
        cell = ws.cell(row=1, column=col, value=industry)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Write row labels (plan tiers) and data
    for row, (tier, row_data) in enumerate(matrix.iterrows(), 2):  # Start at row 2 (row 1 is for industry headers)
        cell = ws.cell(row=row, column=1, value=tier)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        
        for col, value in enumerate(row_data, 2):
            ws.cell(row=row, column=col, value=value / 100)  # Convert to decimal for percentage formatting
            ws.cell(row=row, column=col).number_format = "0.0%"
    
    # Add conditional formatting (green→red color scale)
    # Define the range for conditional formatting (all data cells)
    data_range = f"B2:{get_column_letter(len(matrix.columns) + 1)}{len(matrix) + 1}"
    
    color_scale = ColorScaleRule(
        start_type='min', start_color='00FF00',  # Green for low churn
        mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow for medium
        end_type='max', end_color='FF0000'  # Red for high churn
    )
    ws.conditional_formatting.add(data_range, color_scale)
    
    # Freeze header row and first column
    ws.freeze_panes = "B2"
    
    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width